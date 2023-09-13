import time
import openpyxl

import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.common.by import By

options = uc.ChromeOptions()
driver = uc.Chrome(options=options)

# Создаем новый файл Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
#
# # Задаем заголовки для столбцов
sheet['A1'] = "Название товара"
sheet['B1'] = "Цена после скидок"
sheet['C1'] = "Цена при оплате online"
sheet['D1'] = "Вес изделия"
sheet['E1'] = "Ссылка на товар"
#
row = 2
driver.get("https://www.585zolotoy.ru/catalog/gold-bracelets-with-stones/1/")
time.sleep(5)
prev_scroll_position = 0
while True:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(3)
    current_scroll_position = driver.execute_script("return window.pageYOffset;")
    if current_scroll_position == prev_scroll_position:
        break
    prev_scroll_position = current_scroll_position
print("Начался парс")
#
# # Находим элемент <ul> с классом 'tiles'
ul_element = driver.find_element(By.CLASS_NAME, "tiles")
#Находим все элементы <li> внутри <ul>
li_elements = ul_element.find_elements(By.CSS_SELECTOR, 'li')

# #Проходим по каждому элементу <li>
for li_element in li_elements:
     # Получаем текст
     text = li_element.find_element(By.CLASS_NAME, "product-name").text

# #Получаем ссылку
     link_text = li_element.find_element(By.CLASS_NAME, "product-name")
     link = link_text.find_element(By.TAG_NAME, "a")
     href_value = link.get_attribute("href")

     # Получаем актуальную цену
     actual_price = li_element.find_element(By.CLASS_NAME, "actual-price-row").text

     sheet[f'A{row}'] = text
     sheet[f'B{row}'] = actual_price
     sheet[f'C{row}'] = ""
     sheet[f'D{row}'] = ""
     sheet[f'E{row}'] = href_value
     row +=1
#Сохраняем файл Excel
workbook.save("output.xlsx")
print("Спарсилось")
time.sleep(3)


workbook = openpyxl.load_workbook("output.xlsx")
sheet = workbook.active
# Находим максимальную строку с данными
max_row = sheet.max_row
# Создаем список для хранения ссылок
product_links = []

# Итерируемся по строкам и извлекаем ссылки из столбца "E"
for row in range(2, max_row + 1):
    link = sheet[f'E{row}'].value
    if link:
        product_links.append(link)

print(product_links)
row_number = 2
# Можете перебрать их и выполнить нужные действия, например:
for link in product_links:
    print("\n")
    driver.get(link)
    print(link)
    time.sleep(3)
    try:
        online_price = driver.find_element(By.CLASS_NAME, "online-line")
        price = online_price.find_element(By.CLASS_NAME, "amount").text
    except: price = "Не найдено информации"
    #print(price)

    # Находим элемент <ul> с классом 'features-list'
    ul_element = driver.find_element(By.CLASS_NAME, "features-list")
    # Находим все элементы <li> внутри <ul>
    li_elements = ul_element.find_elements(By.TAG_NAME, "li")
    # Проходим по каждому <li> элементу и ищем тот, в котором есть слово "вес"
    target_li = None
    for li_element in li_elements:
        if "вес" in li_element.text.lower():  # Проверяем, содержит ли текст "вес"
            target_li = li_element
            break
    # Если нашли нужный <li> элемент, извлекаем информацию из него
    if target_li:
        info_from_target_li = target_li.text
        parts = info_from_target_li.split("вес")

        # Взять вторую часть строки (после ключевого слова "вес")
        if len(parts) > 1:
            info_about_weight = parts[1].strip()  # Удаляем лишние пробелы
            print(info_about_weight)
        else:
            print("Ключевое слово 'вес' не найдено в строке")
    else:
        print("Не найдено <li> с информацией о весе")

    sheet.cell(row=row_number, column=3, value=price)
    sheet.cell(row=row_number, column=4, value=info_about_weight)
    row_number += 1

workbook.save("output.xlsx")
time.sleep(5)
driver.quit()
